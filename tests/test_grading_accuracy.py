"""
Tests Markly's grading accuracy against human scores from the AutomatedDataset.

Uses:
  - MIS415Set2  → answer key (question + mark guide + max score)
  - MIS221Set1  → student answers + human-assigned scores

Run:
    PYTHONPATH=. python tests/test_grading_accuracy.py
"""
import logging
import openpyxl

from workers.grader_worker import grade_written

logging.basicConfig(level=logging.WARNING)  # suppress API noise during batch run

XLSX_PATH = "samples/AutomatedDataset.xlsx"

# Questions to test — picked for clear rubrics and good volume of student answers
TEST_QUESTIONS = ["1ai", "2a", "5a"]


def load_answer_key(wb, question_numbers: list[str]) -> dict:
    ws = wb["MIS415Set2"]
    key = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        q_num = row[5]
        if q_num in question_numbers:
            key[q_num] = {
                "number": q_num,
                "type": "written",
                "answer": row[6],
                "rubric": row[7],
                "marks": int(row[8]),
                "score_step": 0.5,      # this dataset uses half-mark increments
            }
    return key


def load_student_answers(wb, question_numbers: list[str]) -> dict[str, list[dict]]:
    ws = wb["MIS221Set1"]
    answers: dict[str, list[dict]] = {q: [] for q in question_numbers}
    for row in ws.iter_rows(min_row=2, values_only=True):
        q_num = row[5]
        if q_num in question_numbers and row[6] and row[7] is not None:
            answers[q_num].append({
                "text": str(row[6]),
                "human_score": float(row[7]),
            })
    return answers


def run_accuracy_test(question_numbers: list[str], sample_size: int = 5):
    wb = openpyxl.load_workbook(XLSX_PATH)
    answer_key = load_answer_key(wb, question_numbers)
    student_answers = load_student_answers(wb, question_numbers)

    overall_errors = []

    for q_num in question_numbers:
        if q_num not in answer_key:
            print(f"Q{q_num}: no answer key entry found — skipping.\n")
            continue

        key_entry = answer_key[q_num]
        answers = student_answers.get(q_num, [])[:sample_size]

        if not answers:
            print(f"Q{q_num}: no student answers found — skipping.\n")
            continue

        max_score = key_entry["marks"]
        print(f"{'='*60}")
        print(f"Q{q_num}  (max {max_score} marks)")
        print(f"Rubric: {key_entry['rubric'][:100]}...")
        print(f"{'='*60}")

        q_errors = []

        for i, ans in enumerate(answers, start=1):
            result = grade_written(ans["text"], key_entry)
            markly_score = result["score"]
            human_score  = ans["human_score"]
            error        = abs(markly_score - human_score)
            q_errors.append(error)
            overall_errors.append(error)

            status = "✓" if error <= 0.5 else "~" if error <= 1.0 else "✗"
            print(
                f"  [{status}] Student {i:02d} | "
                f"Human: {human_score:.1f}  Markly: {markly_score}  "
                f"Error: {error:.1f}"
            )
            print(f"        Feedback: {result['feedback'][:90]}")
            print()

        mae = sum(q_errors) / len(q_errors)
        within_half  = sum(1 for e in q_errors if e <= 0.5)
        within_one   = sum(1 for e in q_errors if e <= 1.0)
        print(f"  Q{q_num} summary: MAE={mae:.2f}  "
              f"Within 0.5: {within_half}/{len(q_errors)}  "
              f"Within 1.0: {within_one}/{len(q_errors)}")
        print()

    if overall_errors:
        overall_mae = sum(overall_errors) / len(overall_errors)
        within_half  = sum(1 for e in overall_errors if e <= 0.5)
        within_one   = sum(1 for e in overall_errors if e <= 1.0)
        total = len(overall_errors)
        print(f"{'='*60}")
        print(f"OVERALL  ({total} answers graded)")
        print(f"  Mean Absolute Error : {overall_mae:.2f}")
        print(f"  Within 0.5 marks    : {within_half}/{total} ({within_half/total*100:.0f}%)")
        print(f"  Within 1.0 mark     : {within_one}/{total}  ({within_one/total*100:.0f}%)")
        print(f"{'='*60}")


if __name__ == "__main__":
    run_accuracy_test(TEST_QUESTIONS, sample_size=5)
