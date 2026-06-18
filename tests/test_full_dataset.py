"""
Full dataset accuracy test against MIS221 student answers.

Uses Gemini first (500 req/day free), automatically falls back to Groq when
Gemini hits its daily limit. Saves results to CSV incrementally so the run
can be resumed if interrupted.

Usage:
    PYTHONPATH=. python tests/test_full_dataset.py

Results saved to: samples/dataset_results.csv
"""
import csv
import logging
import os
import time
import openpyxl

from workers.grader_worker import grade_written

logging.basicConfig(level=logging.WARNING)

XLSX_PATH    = "samples/AutomatedDataset.xlsx"
RESULTS_CSV  = "samples/dataset_results.csv"
RATE_LIMIT_DELAY = 2.1  # seconds between calls — keeps Groq well under 30 RPM

CSV_HEADERS = [
    "question", "student_index", "human_score", "markly_score",
    "error", "max_score", "feedback",
]


def load_answer_key(wb) -> dict:
    ws = wb["MIS415Set2"]
    key = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        q_num, question, mark_guide, max_score = row[5], row[6], row[7], row[8]
        if q_num and mark_guide and max_score:
            key[q_num] = {
                "number": q_num,
                "type": "written",
                "answer": str(question) if question else "",
                "rubric": str(mark_guide),
                "marks": int(max_score),
                "score_step": 0.5,      # this dataset uses half-mark increments
            }
    return key


def load_student_answers(wb) -> list[dict]:
    ws = wb["MIS221Set1"]
    answers = []
    counters: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        q_num, text, human_score = row[5], row[6], row[7]
        if q_num and text and human_score is not None:
            counters[q_num] = counters.get(q_num, 0) + 1
            answers.append({
                "question": q_num,
                "student_index": counters[q_num],
                "text": str(text),
                "human_score": float(human_score),
            })
    return answers


def load_already_done() -> set[tuple]:
    """Return (question, student_index) pairs already in the CSV."""
    done = set()
    if not os.path.exists(RESULTS_CSV):
        return done
    with open(RESULTS_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            done.add((row["question"], int(row["student_index"])))
    return done


def append_result(row: dict):
    is_new = not os.path.exists(RESULTS_CSV)
    with open(RESULTS_CSV, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        if is_new:
            writer.writeheader()
        writer.writerow(row)


def print_summary(results: list[dict]):
    if not results:
        print("No results to summarise.")
        return

    errors = [abs(r["markly_score"] - r["human_score"]) for r in results]
    mae = sum(errors) / len(errors)
    within_half = sum(1 for e in errors if e <= 0.5)
    within_one  = sum(1 for e in errors if e <= 1.0)
    total = len(errors)

    # Per-question breakdown
    by_q: dict[str, list] = {}
    for r in results:
        by_q.setdefault(r["question"], []).append(abs(r["markly_score"] - r["human_score"]))

    print(f"\n{'='*64}")
    print(f"FULL DATASET RESULTS  ({total} answers across {len(by_q)} questions)")
    print(f"{'='*64}")
    print(f"  Mean Absolute Error : {mae:.3f}")
    print(f"  Within 0.5 marks    : {within_half}/{total} ({within_half/total*100:.1f}%)")
    print(f"  Within 1.0 mark     : {within_one}/{total}  ({within_one/total*100:.1f}%)")
    print(f"\n  Per-question MAE:")
    for q, errs in sorted(by_q.items()):
        q_mae = sum(errs) / len(errs)
        print(f"    Q{q:<8}  n={len(errs):3d}  MAE={q_mae:.2f}")
    print(f"{'='*64}")


def run():
    wb = openpyxl.load_workbook(XLSX_PATH)
    answer_key    = load_answer_key(wb)
    student_answers = load_student_answers(wb)
    already_done  = load_already_done()

    to_grade = [
        a for a in student_answers
        if a["question"] in answer_key
        and (a["question"], a["student_index"]) not in already_done
    ]

    print(f"Total answers to grade: {len(to_grade)}  (skipping {len(already_done)} already done)")
    print(f"Estimated time at {RATE_LIMIT_DELAY}s/call: {len(to_grade) * RATE_LIMIT_DELAY / 60:.0f} min\n")

    new_results = []

    for i, ans in enumerate(to_grade, start=1):
        q_num      = ans["question"]
        key_entry  = answer_key[q_num]
        human_score = ans["human_score"]

        try:
            result = grade_written(ans["text"], key_entry)
            markly_score = result["score"]
            feedback     = result.get("feedback", "")[:120]
        except Exception as e:
            print(f"  ERROR on Q{q_num} student {ans['student_index']}: {e}")
            markly_score = -1
            feedback     = f"ERROR: {e}"

        error = abs(markly_score - human_score) if markly_score >= 0 else None
        row = {
            "question":      q_num,
            "student_index": ans["student_index"],
            "human_score":   human_score,
            "markly_score":  markly_score,
            "error":         round(error, 2) if error is not None else "ERROR",
            "max_score":     key_entry["marks"],
            "feedback":      feedback,
        }
        append_result(row)
        new_results.append({**ans, "markly_score": markly_score, "human_score": human_score})

        if i % 10 == 0 or i == len(to_grade):
            done_errors = [abs(r["markly_score"] - r["human_score"]) for r in new_results if r["markly_score"] >= 0]
            mae_so_far = sum(done_errors) / len(done_errors) if done_errors else 0
            print(f"  [{i:4d}/{len(to_grade)}]  Q{q_num} s{ans['student_index']:03d}  "
                  f"human={human_score:.1f}  markly={markly_score}  "
                  f"running MAE={mae_so_far:.2f}")

        time.sleep(RATE_LIMIT_DELAY)

    # Load all results (including prior runs) for final summary
    all_results = []
    with open(RESULTS_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row["error"] != "ERROR":
                all_results.append({
                    "question":      row["question"],
                    "human_score":   float(row["human_score"]),
                    "markly_score":  float(row["markly_score"]),
                })

    print_summary(all_results)
    print(f"\nFull results saved to: {RESULTS_CSV}")


if __name__ == "__main__":
    run()
